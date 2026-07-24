from functools import wraps
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_file
import data_store
import config
import db
from classifier import classify as _classify
import pid_kb as _kb

app = Flask(__name__)
app.secret_key = config.PORTAL_SECRET_KEY or "dev-only-change-me"

STATUSES = ["Em Aberto", "Em Análise", "Aprovada", "Rejeitada", "Ganha", "Perdida"]
FORECAST_STATUSES = data_store.FORECAST_STATUSES

_IMPORT_COLS = (
    "id", "subject", "request_type", "project_type", "project_ref",
    "requester_name", "department", "cnpj",
    "smart_account", "smart_account_domain", "virtual_account",
    "status", "valor_total", "responsavel_interno", "fornecedor", "observacoes",
    "projeto_id_vale", "logicalis_id", "ntt_id",
    "estimate_nacional", "estimate_importado", "order_id", "deal_id",
    "date", "response_received_at",
)
_IMPORT_COL_WIDTHS = (
    12, 42, 18, 22, 15, 24, 22, 22, 18, 24, 18,
    20, 16, 24, 15, 32, 18, 18, 15, 18, 20, 15, 15,
    18, 20,
)

# 'date' guarda 'YYYY-MM-DD[ HH:MM:SS]' internamente; aceitamos também DD/MM/YYYY
# na planilha (formato mais comum de digitação manual no Brasil).
_IMPORT_DATE_FORMATS = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y")


def _parse_import_date(raw: str) -> str:
    """Normaliza uma data vinda da planilha (célula de data do Excel já vem
    como 'YYYY-MM-DD...' via str(); texto solto pode vir em DD/MM/YYYY).
    Retorna 'YYYY-MM-DD HH:MM:SS', ou '' se vazio/não reconhecido (nunca
    lança — planilha com data inválida não deve travar a linha inteira)."""
    raw = (raw or "").strip()
    if not raw:
        return ""
    for fmt in _IMPORT_DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return ""


@app.template_filter("date_br")
def date_br(value):
    parts = str(value).split("-")
    if len(parts) == 3:
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    return value


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        if session.get("role") not in ("admin", "owner"):
            flash("Acesso restrito a administradores.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated


def owner_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        if session.get("role") != "owner":
            return jsonify({"ok": False, "error": "Acesso restrito ao Owner."}), 403
        return f(*args, **kwargs)
    return decorated


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if "user" in session:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        u = request.form.get("username", "").strip()
        p = request.form.get("password", "")
        u_data = data_store.verify_user(u, p)
        if u_data:
            session["user"] = u
            session["role"] = u_data.get("role", "viewer")
            session["empresa"] = u_data.get("empresa", "")
            return redirect(url_for("dashboard"))
        flash("Usuário ou senha incorretos.", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/profile/password", methods=["POST"])
@login_required
def profile_password():
    current  = request.form.get("current_password", "")
    new_pwd  = request.form.get("new_password", "")
    confirm  = request.form.get("confirm_password", "")
    if not current or not new_pwd:
        return jsonify({"ok": False, "error": "Preencha todos os campos."}), 400
    if new_pwd != confirm:
        return jsonify({"ok": False, "error": "As senhas não coincidem."}), 400
    if len(new_pwd) < 6:
        return jsonify({"ok": False, "error": "Senha deve ter no mínimo 6 caracteres."}), 400
    ok, err = data_store.change_password(session["user"], current, new_pwd)
    if not ok:
        return jsonify({"ok": False, "error": err}), 400
    return jsonify({"ok": True})


# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def dashboard():
    return render_template("dashboard.html", stats=data_store.get_stats())


@app.route("/api/stats")
@login_required
def api_stats():
    return jsonify(data_store.get_stats())


@app.route("/api/forecast")
@login_required
def api_forecast():
    after = request.args.get("after", "")
    return jsonify(data_store.get_delivery_forecasts(after))


# ── Quotes ────────────────────────────────────────────────────────────────────

@app.route("/quotes")
@login_required
def quotes():
    all_quotes  = data_store.load_extractions()
    annotations = data_store.load_annotations()
    timelines   = data_store.load_timelines()

    tipo_f       = request.args.get("tipo", "")
    status_f     = request.args.get("status", "")
    tl_step_f    = request.args.get("tl_step", "")
    fornecedor_f = request.args.get("fornecedor", "")
    search       = request.args.get("q", "").lower()

    all_fornecedores = sorted({
        ann.get("fornecedor", "").strip()
        for ann in annotations.values() if ann.get("fornecedor", "").strip()
    })

    result = []
    for q in all_quotes:
        ann = annotations.get(q["id"], {})
        q["_status"]      = ann.get("status", "Em Aberto")
        q["_valor"]       = ann.get("valor_total") or ""
        q["_resp"]        = ann.get("responsavel_interno", "")
        q["_fornecedor"]  = ann.get("fornecedor", "")

        # Compute current timeline step
        request_type = q.get("request_type", "Cotação")
        tl_steps     = data_store.TIMELINE_STEPS.get(request_type, data_store.TIMELINE_STEPS["Cotação"])
        tl_dates     = timelines.get(q["id"], {}).get("dates", {})
        done_count   = sum(1 for s in tl_steps if tl_dates.get(s["key"]))
        total_steps  = len(tl_steps)
        if done_count == total_steps:
            q["_tl_label"] = "Concluído"
            q["_tl_icon"]  = "fa-flag-checkered"
            q["_tl_pct"]   = 100
        elif done_count == 0:
            q["_tl_label"] = tl_steps[0]["label"]
            q["_tl_icon"]  = tl_steps[0]["icon"]
            q["_tl_pct"]   = 0
        else:
            current        = tl_steps[done_count]
            q["_tl_label"] = current["label"]
            q["_tl_icon"]  = current["icon"]
            q["_tl_pct"]   = int(done_count / total_steps * 100)
        q["_tl_done"]  = done_count
        q["_tl_total"] = total_steps

        if tipo_f    and q.get("request_type") != tipo_f:
            continue
        if status_f  and q["_status"] != status_f:
            continue
        if tl_step_f and q["_tl_label"] != tl_step_f:
            continue
        if fornecedor_f and q["_fornecedor"] != fornecedor_f:
            continue
        if search:
            haystack = " ".join([
                q.get("subject", ""), q.get("from", ""),
                q.get("requester_name", ""), q.get("department", ""),
            ]).lower()
            if search not in haystack:
                continue
        result.append(q)

    # Collect unique tl_step labels for filter dropdown
    all_tl_labels = sorted({q["_tl_label"] for q in result})

    return render_template(
        "quotes.html", quotes=result,
        tipo_f=tipo_f, status_f=status_f, tl_step_f=tl_step_f, fornecedor_f=fornecedor_f, search=search,
        statuses=STATUSES, tl_step_labels=all_tl_labels, fornecedores=all_fornecedores,
        is_owner=(session.get("role") == "owner"),
    )


@app.route("/quotes/delete", methods=["POST"])
@owner_required
def quotes_delete():
    ids = request.json.get("ids", [])
    if not ids:
        return jsonify({"ok": False, "error": "Nenhuma selecionada"}), 400
    count = data_store.delete_quotes(ids, session["user"])
    return jsonify({"ok": True, "deleted": count})


@app.route("/quotes/<quote_id>")
@login_required
def quote_detail(quote_id):
    all_quotes = data_store.load_extractions()
    idx = next((i for i, q in enumerate(all_quotes) if q["id"] == quote_id), None)
    if idx is None:
        return "Cotação não encontrada", 404
    quote     = all_quotes[idx]
    prev_id   = all_quotes[idx - 1]["id"] if idx > 0 else None
    next_id   = all_quotes[idx + 1]["id"] if idx < len(all_quotes) - 1 else None
    ann      = data_store.load_annotations().get(quote_id, {})
    corr     = data_store.load_corrections().get(quote_id, {})
    timeline = data_store.load_timelines().get(quote_id, {"dates": {}})
    deal     = data_store.load_deals().get(quote_id, {})
    request_type = corr.get("request_type", {}).get("current") or quote.get("request_type", "Cotação")
    tl_steps = data_store.TIMELINE_STEPS.get(request_type, data_store.TIMELINE_STEPS["Cotação"])

    # ── Cálculo automático de previsão de conclusão ──
    auto_forecast, aggressor = data_store.compute_auto_forecast(quote, timeline, deal)

    return render_template("quote_detail.html", quote=quote, ann=ann, corr=corr,
                           statuses=STATUSES, correctable=data_store.CORRECTABLE_FIELDS,
                           timeline=timeline, timeline_steps=tl_steps,
                           deal=deal, deal_fields=data_store.DEAL_FIELDS,
                           auto_forecast=auto_forecast, aggressor=aggressor,
                           prev_id=prev_id, next_id=next_id)


@app.route("/quotes/new", methods=["GET", "POST"])
@login_required
def quote_new():
    if request.method == "POST":
        import uuid
        quote_id = str(uuid.uuid4())
        now_str  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Produtos dinâmicos
        products = []
        for qty, part, desc in zip(
            request.form.getlist("qty[]"),
            request.form.getlist("part_number[]"),
            request.form.getlist("description[]"),
        ):
            part = part.strip()
            if part:
                products.append({"qty": qty.strip() or "1", "part_number": part, "description": desc.strip()})

        quote = {
            "id":                   quote_id,
            "is_manual":            True,
            "date":                 now_str,
            "from":                 session["user"],
            "subject":              request.form.get("subject", "").strip() or "(sem assunto)",
            "request_type":         request.form.get("request_type", "Cotação"),
            "project_type":         request.form.get("project_type", "NA"),
            "requester_name":       request.form.get("requester_name", "").strip(),
            "department":           request.form.get("department", "").strip(),
            "recipient":            config.EMAIL_ADDRESS,
            "cnpj":                 request.form.get("cnpj", "").strip(),
            "smart_account":        request.form.get("smart_account", "").strip(),
            "smart_account_domain": request.form.get("smart_account_domain", "").strip(),
            "virtual_account":      request.form.get("virtual_account", "").strip(),
            "project_ref":          request.form.get("project_ref", "").strip() or "NA",
            "products":             products,
            "body":                 request.form.get("body", "").strip(),
        }

        data_store._insert_extraction(quote)

        raw_valor = request.form.get("valor_total", "").replace(",", ".").replace("R$", "").strip()
        ann_data  = {
            "status":              request.form.get("status", "Em Aberto"),
            "valor_total":         float(raw_valor) if raw_valor else None,
            "responsavel_interno": request.form.get("responsavel_interno", "").strip(),
            "fornecedor":          request.form.get("fornecedor", "").strip(),
            "observacoes":         request.form.get("observacoes", "").strip(),
        }
        data_store.save_annotation(quote_id, quote["subject"], ann_data, session["user"])

        flash("Cotação criada com sucesso!", "success")
        return redirect(url_for("quote_detail", quote_id=quote_id))

    return render_template("new_quote.html", statuses=STATUSES,
                           now=datetime.now().strftime("%Y-%m-%dT%H:%M"))


@app.route("/quotes/<quote_id>/products", methods=["POST"])
@login_required
def quote_products(quote_id):
    quote = next((q for q in data_store.load_extractions() if q["id"] == quote_id), None)
    if not quote:
        return "Cotação não encontrada", 404
    products = []
    fields = zip(
        request.form.getlist("qty[]"),
        request.form.getlist("part_number[]"),
        request.form.getlist("description[]"),
        request.form.getlist("unit_list_price[]"),
        request.form.getlist("lead_time[]"),
        request.form.getlist("discount_pct[]"),
        request.form.getlist("unit_net_price[]"),
        request.form.getlist("extended_net_price[]"),
        request.form.getlist("tipo[]"),
        request.form.getlist("arquitetura[]"),
        request.form.getlist("categoria[]"),
    )
    for qty, part, desc, ulp, lt, disc, unp, enp, tipo, arq, cat in fields:
        part = part.strip()
        if part:
            resolved_arch, resolved_cat = _kb.classify_with_kb(part, desc.strip(), _classify)
            products.append({
                "qty":                qty.strip() or "1",
                "part_number":        part,
                "description":        desc.strip(),
                "unit_list_price":    ulp.strip(),
                "lead_time":          lt.strip(),
                "discount_pct":       disc.strip(),
                "unit_net_price":     unp.strip(),
                "extended_net_price": enp.strip(),
                "tipo":               tipo or "Nacional",
                "arquitetura":        arq or resolved_arch,
                "categoria":          cat or resolved_cat,
            })
    # Detect manual overrides and persist to KB
    old_products = quote.get("products", [])
    corrections = _kb.apply_manual_corrections(old_products, products)
    data_store.save_products(quote_id, quote.get("subject", ""), products, session["user"])
    msg = f"Produtos atualizados! {corrections} correção(ões) salva(s) na base de PIDs." if corrections else "Produtos atualizados com sucesso!"
    flash(msg, "success")
    return redirect(url_for("quote_detail", quote_id=quote_id))


@app.route("/quotes/<quote_id>/products/upload", methods=["POST"])
@login_required
def quote_products_upload(quote_id):
    quote = next((q for q in data_store.load_extractions() if q["id"] == quote_id), None)
    if not quote:
        return "Cotação não encontrada", 404

    file = request.files.get("xls_file")
    if not file or not file.filename:
        flash("Nenhum arquivo enviado.", "danger")
        return redirect(url_for("quote_detail", quote_id=quote_id))

    fname = file.filename.lower()
    if not fname.endswith((".xls", ".xlsx")):
        flash("Formato inválido. Envie um arquivo .xls ou .xlsx exportado do CCW.", "danger")
        return redirect(url_for("quote_detail", quote_id=quote_id))

    tipo = request.form.get("tipo", "Nacional")
    buf = file.read()
    products = []

    def _fmt(v):
        if v is None or v == "":
            return ""
        if isinstance(v, float):
            return str(int(v)) if v == int(v) else f"{v:.2f}"
        return str(v).strip()

    # Detect column indices dynamically from the header row.
    # Supports the modelo.xlsx layout, legacy CCW exports (Part Number in col A)
    # and the CCW "Order Status" export (Item Name / Order Line Details tab).
    _COL_KEYS = {
        "part number":        "part",
        "item name":          "part",
        "description":        "desc",
        "qty":                "qty",
        "quantity":           "qty",
        "unit list price":    "ulp",
        "list price":         "ulp",
        "estimated lead time": "lt",
        "disc(%)":            "disc",
        "discount":           "disc",
        "unit net price":     "unp",
        "extended net price": "enp",
    }
    _HEADER_HINTS = ("part number", "item name")

    def _find_cols(header_cells):
        cols = {}
        for idx, cell in enumerate(header_cells):
            key = str(cell or "").strip().lower()
            for pattern, name in _COL_KEYS.items():
                if pattern in key and name not in cols:
                    cols[name] = idx
        return cols

    def _is_header_row(header_cells):
        return any(hint in str(c or "").lower() for c in header_cells for hint in _HEADER_HINTS)

    try:
        # Content sniffing: the real .xlsx (zip) signature takes precedence over the
        # file extension, since CCW "Order Status" exports are sometimes downloaded
        # with a .xls extension while actually containing .xlsx (zip) data.
        if buf[:4] != b"PK\x03\x04":
            import xlrd
            wb = xlrd.open_workbook(file_contents=buf)
            ws = hdr_row = None
            for sheet in wb.sheets():
                for r in range(sheet.nrows):
                    if _is_header_row([sheet.cell_value(r, c) for c in range(sheet.ncols)]):
                        ws, hdr_row = sheet, r
                        break
                if ws:
                    break
            if ws is None:
                flash("Coluna 'Part Number' não encontrada — verifique se é um export CCW.", "danger")
                return redirect(url_for("quote_detail", quote_id=quote_id))
            cols = _find_cols([ws.cell_value(hdr_row, c) for c in range(ws.ncols)])
            for r in range(hdr_row + 1, ws.nrows):
                part = str(ws.cell_value(r, cols["part"])).strip() if "part" in cols else ""
                if not part or part.lower() in ("none", "nan") or " " in part or ":" in part:
                    continue
                desc = _fmt(ws.cell_value(r, cols["desc"])) if "desc" in cols else ""
                arch, cat = _kb.classify_with_kb(part, desc, _classify)
                products.append({
                    "qty":                _fmt(ws.cell_value(r, cols["qty"]))  if "qty"  in cols else "1",
                    "part_number":        part,
                    "description":        desc,
                    "unit_list_price":    _fmt(ws.cell_value(r, cols["ulp"]))  if "ulp"  in cols else "",
                    "lead_time":          _fmt(ws.cell_value(r, cols["lt"]))   if "lt"   in cols else "",
                    "discount_pct":       _fmt(ws.cell_value(r, cols["disc"])) if "disc" in cols else "",
                    "unit_net_price":     _fmt(ws.cell_value(r, cols["unp"]))  if "unp"  in cols else "",
                    "extended_net_price": _fmt(ws.cell_value(r, cols["enp"]))  if "enp"  in cols else "",
                    "tipo":               tipo,
                    "arquitetura":        arch,
                    "categoria":          cat,
                })
        else:
            from openpyxl import load_workbook
            from io import BytesIO
            wb = load_workbook(BytesIO(buf), data_only=True)
            ws = hdr_row = None
            for sheet in wb.worksheets:
                if sheet.sheet_state != "visible":
                    continue
                for r in range(1, sheet.max_row + 1):
                    if _is_header_row([sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]):
                        ws, hdr_row = sheet, r
                        break
                if ws:
                    break
            if ws is None:
                flash("Coluna 'Part Number' não encontrada — verifique se é um export CCW.", "danger")
                return redirect(url_for("quote_detail", quote_id=quote_id))
            cols = _find_cols([ws.cell(hdr_row, c).value for c in range(1, ws.max_column + 1)])

            # Order-level exports (e.g. "Order Line Details" tab) don't carry a
            # per-line lead time — fall back to the order's overall estimate.
            fallback_lt = ""
            if "lt" not in cols and "Order Header" in wb.sheetnames:
                oh = wb["Order Header"]
                for r in range(1, oh.max_row + 1):
                    if "estimated lead time" in str(oh.cell(r, 1).value or "").lower():
                        fallback_lt = _fmt(oh.cell(r, 2).value)
                        break

            for r in range(hdr_row + 1, ws.max_row + 1):
                get = lambda k: ws.cell(r, cols[k] + 1).value if k in cols else None
                part = str(get("part") or "").strip()
                if not part or part.lower() in ("none", "nan") or " " in part or ":" in part:
                    continue
                desc = _fmt(get("desc"))
                arch, cat = _kb.classify_with_kb(part, desc, _classify)
                products.append({
                    "qty":                _fmt(get("qty"))  or "1",
                    "part_number":        part,
                    "description":        desc,
                    "unit_list_price":    _fmt(get("ulp")),
                    "lead_time":          _fmt(get("lt")) if "lt" in cols else fallback_lt,
                    "discount_pct":       _fmt(get("disc")),
                    "unit_net_price":     _fmt(get("unp")),
                    "extended_net_price": _fmt(get("enp")),
                    "tipo":               tipo,
                    "arquitetura":        arch,
                    "categoria":          cat,
                })
    except Exception as exc:
        flash(f"Erro ao processar arquivo: {exc}", "danger")
        return redirect(url_for("quote_detail", quote_id=quote_id))

    if not products:
        flash("Nenhum produto encontrado no arquivo.", "warning")
        return redirect(url_for("quote_detail", quote_id=quote_id))

    data_store.append_products(quote_id, quote.get("subject", ""), products, session["user"])
    flash(f"{len(products)} produto(s) importado(s) ({tipo}) e adicionados à lista!", "success")
    return redirect(url_for("quote_detail", quote_id=quote_id))


@app.route("/quotes/<quote_id>/deal", methods=["POST"])
@login_required
def quote_deal(quote_id):
    quote = next((q for q in data_store.load_extractions() if q["id"] == quote_id), None)
    if not quote:
        return "Cotação não encontrada", 404
    data = {f["key"]: request.form.get(f["key"], "").strip() for f in data_store.DEAL_FIELDS}
    data_store.save_deal(quote_id, quote.get("subject", ""), data, session["user"])
    flash("IDs e Estimates salvos com sucesso!", "success")
    return redirect(url_for("quote_detail", quote_id=quote_id))


@app.route("/quotes/<quote_id>/timeline", methods=["POST"])
@login_required
def quote_timeline(quote_id):
    quote = next((q for q in data_store.load_extractions() if q["id"] == quote_id), None)
    if not quote:
        return "Cotação não encontrada", 404
    corr = data_store.load_corrections().get(quote_id, {})
    request_type = corr.get("request_type", {}).get("current") or quote.get("request_type", "Cotação")
    steps = data_store.TIMELINE_STEPS.get(request_type, data_store.TIMELINE_STEPS["Cotação"])
    dates = {}
    for step in steps:
        val = request.form.get(step["key"], "").strip()
        if val:
            dates[step["key"]] = val
        if "extra_key" in step:
            extra_val = request.form.get(step["extra_key"], "").strip()
            if extra_val:
                dates[step["extra_key"]] = extra_val
    data_store.save_timeline(quote_id, quote.get("subject", ""), dates, session["user"])
    flash("Timeline atualizada com sucesso!", "success")
    return redirect(url_for("quote_detail", quote_id=quote_id))


@app.route("/quotes/<quote_id>/correct", methods=["POST"])
@login_required
def quote_correct(quote_id):
    quote = next((q for q in data_store.load_extractions() if q["id"] == quote_id), None)
    if not quote:
        return "Cotação não encontrada", 404
    fields = {f: request.form.get(f, "").strip() for f in data_store.CORRECTABLE_FIELDS}
    data_store.save_correction(quote_id, quote.get("subject", ""), fields, session["user"])
    flash("Dados extraídos corrigidos com sucesso!", "success")
    return redirect(url_for("quote_detail", quote_id=quote_id))


@app.route("/quotes/<quote_id>/inline", methods=["POST"])
@login_required
def quote_inline_edit(quote_id):
    """Salva um único campo da cotação via fetch (JSON). Usado pela edição inline da lista."""
    quote = next((q for q in data_store.load_extractions() if q["id"] == quote_id), None)
    if not quote:
        return jsonify({"ok": False, "error": "não encontrada"}), 404
    data = request.get_json(silent=True) or {}
    field = data.get("field", "")
    value = str(data.get("value", "")).strip()
    allowed = {"subject", "requester_name", "department", "request_type"}
    if field not in allowed or not value:
        return jsonify({"ok": False, "error": "campo inválido"}), 400
    data_store.update_extraction_fields(quote_id, {field: value}, quote.get("subject", ""), session["user"])
    return jsonify({"ok": True, "value": value})


@app.route("/quotes/<quote_id>/edit", methods=["POST"])
@login_required
def quote_edit(quote_id):
    quote = next((q for q in data_store.load_extractions() if q["id"] == quote_id), None)
    if not quote:
        return "Cotação não encontrada", 404

    raw_valor = request.form.get("valor_total", "").replace(",", ".").replace("R$", "").strip()
    new_data = {
        "valor_total":        float(raw_valor) if raw_valor else None,
        "status":             request.form.get("status", "Em Aberto"),
        "responsavel_interno": request.form.get("responsavel_interno", "").strip(),
        "fornecedor":         request.form.get("fornecedor", "").strip(),
        "observacoes":        request.form.get("observacoes", "").strip(),
        "entregue":           request.form.get("entregue") == "1",
    }
    data_store.save_annotation(quote_id, quote.get("subject", ""), new_data, session["user"])
    # Reflete o checkbox "Entregue" (ou um prazo já vencido) na etapa imediatamente,
    # sem esperar o próximo sync do bot de CCW.
    data_store.check_and_advance_by_deadline(quote_id, quote.get("subject", ""), session["user"])
    flash("Cotação atualizada com sucesso!", "success")
    return redirect(url_for("quote_detail", quote_id=quote_id))


# ── Admin ─────────────────────────────────────────────────────────────────────

@app.route("/admin")
@admin_required
def admin():
    return render_template("admin.html", users=data_store.list_users_safe())


@app.route("/admin/users", methods=["POST"])
@admin_required
def admin_create_user():
    username  = request.form.get("username", "").strip()
    password  = request.form.get("password", "")
    password2 = request.form.get("password_confirm", "")
    nome      = request.form.get("nome", "").strip()
    email     = request.form.get("email", "").strip()
    celular   = request.form.get("celular", "").strip()
    empresa   = request.form.get("empresa", "").strip()
    role      = request.form.get("role", "viewer")

    if not username or not password:
        flash("Username e senha são obrigatórios.", "danger")
        return redirect(url_for("admin"))
    if password != password2:
        flash("As senhas não coincidem.", "danger")
        return redirect(url_for("admin"))
    if data_store.username_exists(username):
        flash(f"Username '{username}' já existe.", "danger")
        return redirect(url_for("admin"))

    data_store.create_user(username, password, role=role,
                           nome=nome, email=email, celular=celular, empresa=empresa)
    flash(f"Usuário '{username}' criado com sucesso!", "success")
    return redirect(url_for("admin"))


@app.route("/admin/users/<username>/delete", methods=["POST"])
@admin_required
def admin_delete_user(username):
    if session.get("role") != "owner":
        flash("Apenas o Owner pode deletar usuários.", "danger")
        return redirect(url_for("admin"))
    if username == session.get("user"):
        flash("Você não pode deletar sua própria conta.", "warning")
        return redirect(url_for("admin"))
    if data_store.delete_user(username):
        flash(f"Usuário '{username}' deletado com sucesso.", "success")
    else:
        flash(f"Usuário '{username}' não encontrado.", "danger")
    return redirect(url_for("admin"))


@app.route("/cisco")
@login_required
def cisco():
    if session.get("empresa") != "Cisco":
        flash("Acesso restrito a usuários Cisco.", "danger")
        return redirect(url_for("dashboard"))
    forecast_data = data_store.get_sales_forecast()
    return render_template(
        "cisco.html",
        stats=data_store.get_cisco_spend_stats(),
        discount_history=data_store.get_discount_history(),
        sales_forecast=forecast_data["items"],
        fx_rate=forecast_data["fx_rate"],
        fx_rate_is_live=forecast_data["fx_rate_is_live"],
        forecast_statuses=FORECAST_STATUSES,
    )


@app.route("/cisco/forecast/<quote_id>", methods=["POST"])
@login_required
def cisco_forecast_save(quote_id):
    if session.get("empresa") != "Cisco":
        return "Acesso restrito a usuários Cisco.", 403
    quote = next((q for q in data_store.load_extractions() if q["id"] == quote_id), None)
    if not quote:
        return "Cotação não encontrada", 404

    data = {
        "booking_date": request.form.get("booking_date", "").strip(),
        "tech_lead":    request.form.get("tech_lead", "").strip(),
        "pm_name":      request.form.get("pm_name", "").strip(),
        "status":       request.form.get("status", "Pipeline").strip(),
    }
    data_store.save_cisco_forecast(quote_id, quote.get("subject", ""), data, session["user"])
    flash("Forecast atualizado.", "success")
    return redirect(url_for("cisco") + "#tab-forecast")


@app.route("/admin/import")
@admin_required
def admin_import():
    result = session.pop("import_result", None)
    return render_template("admin_import.html", result=result)


@app.route("/admin/import/template")
@admin_required
def admin_import_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotações"

    hdr_fill = PatternFill("solid", fgColor="005073")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    ex_fill  = PatternFill("solid", fgColor="E8F4F8")
    ex_font  = Font(italic=True, color="888888", size=10)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Header row
    for idx, (key, width) in enumerate(zip(_IMPORT_COLS, _IMPORT_COL_WIDTHS), 1):
        cell = ws.cell(1, idx, key)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "B2"

    # Example row
    example = [
        "",                              # id (vazio = novo)
        "Cotação Cisco Catalyst 9300",   # subject
        "Cotação",                       # request_type
        "Novo Projeto",                  # project_type
        "PROJ-2026-001",                 # project_ref
        "João da Silva",                 # requester_name
        "TI - Infraestrutura",           # department
        "33.592.510/0001-54",            # cnpj
        "BR (Brasil)",                   # smart_account
        "vale.com.br",                   # smart_account_domain
        "BR-TI-SP",                      # virtual_account
        "Em Aberto",                     # status
        "250000",                        # valor_total
        "Maria Fernanda",                # responsavel_interno
        "Logicalis",                     # fornecedor
        "Exemplo de observação",         # observacoes
        "PROJ-VALE-001",                 # projeto_id_vale
        "LOG-12345",                     # logicalis_id
        "",                              # ntt_id
        "280000",                        # estimate_nacional
        "",                              # estimate_importado
        "",                              # order_id
        "",                              # deal_id
        "15/06/2026",                    # date (Data da Solicitação)
        "20/06/2026",                    # response_received_at (Resposta da Cotação)
    ]
    for idx, val in enumerate(example, 1):
        cell = ws.cell(2, idx, val)
        cell.font = ex_font
        cell.fill = ex_fill

    # Dropdown validations
    def _dv(formula, sqref):
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=False)
        ws.add_data_validation(dv)
        dv.sqref = sqref

    _dv('"Cotação,Pedido"',                                               "C3:C10000")
    _dv('"Novo Projeto,Obsolescência,NA"',                                "D3:D10000")
    _dv('"Em Aberto,Em Análise,Aprovada,Rejeitada,Ganha,Perdida"',        "L3:L10000")
    _dv('"Logicalis,NTT,Outros"',                                         "O3:O10000")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="template_cotacoes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/import/upload", methods=["POST"])
@admin_required
def admin_import_upload():
    import uuid as _uuid
    from openpyxl import load_workbook
    from io import BytesIO

    file = request.files.get("file")
    if not file or not file.filename:
        flash("Nenhum arquivo enviado.", "danger")
        return redirect(url_for("admin_import"))

    if not file.filename.lower().endswith((".xlsx", ".xls")):
        flash("Formato inválido. Envie um arquivo .xlsx.", "danger")
        return redirect(url_for("admin_import"))

    try:
        buf = BytesIO(file.read())
        wb  = load_workbook(buf, data_only=True)
        ws  = wb.active
    except Exception as exc:
        flash(f"Erro ao ler o arquivo: {exc}", "danger")
        return redirect(url_for("admin_import"))

    # Validate header row
    actual_hdrs = [
        str(ws.cell(1, c).value or "").strip()
        for c in range(1, len(_IMPORT_COLS) + 1)
    ]
    if actual_hdrs != list(_IMPORT_COLS):
        flash("Cabeçalhos não correspondem ao template. Baixe novamente e use como base.", "danger")
        return redirect(url_for("admin_import"))

    created, updated, errors = 0, 0, []
    user    = session["user"]
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    deal_keys = [f["key"] for f in data_store.DEAL_FIELDS]

    for row_idx in range(2, ws.max_row + 1):
        row = {
            key: (str(ws.cell(row_idx, c).value).strip()
                  if ws.cell(row_idx, c).value is not None else "")
            for c, key in enumerate(_IMPORT_COLS, 1)
        }

        subject = row["subject"]
        if not subject:
            continue  # skip empty / example rows

        quote_id = row["id"]

        extract_fields = {
            "subject":               subject,
            "request_type":          row["request_type"]         or "Cotação",
            "project_type":          row["project_type"]         or "NA",
            "project_ref":           row["project_ref"]          or "NA",
            "requester_name":        row["requester_name"],
            "department":            row["department"],
            "cnpj":                  row["cnpj"],
            "smart_account":         row["smart_account"],
            "smart_account_domain":  row["smart_account_domain"],
            "virtual_account":       row["virtual_account"],
        }

        raw_valor = row["valor_total"].replace(",", ".").replace("R$", "").strip()
        ann_data = {
            "status":               row["status"] or "Em Aberto",
            "valor_total":          float(raw_valor) if raw_valor else None,
            "responsavel_interno":  row["responsavel_interno"],
            "fornecedor":           row["fornecedor"],
            "observacoes":          row["observacoes"],
        }

        deal_data = {k: row.get(k, "") for k in deal_keys}
        deal_clean = {k: v for k, v in deal_data.items() if v}

        parsed_date = _parse_import_date(row.get("date", ""))
        if parsed_date:
            extract_fields["date"] = parsed_date  # só entra se preenchida — nunca apaga a data existente

        resp_date = _parse_import_date(row.get("response_received_at", ""))
        if resp_date:
            deal_clean["response_received_at"] = resp_date

        # Resolve which existing quote to update (if any)
        matched_id = None
        if quote_id:
            if data_store.get_extraction_by_id(quote_id):
                matched_id = quote_id
            else:
                errors.append(f"Linha {row_idx}: ID '{quote_id}' não encontrado.")
                continue
        else:
            # Try deal fields in priority order
            for deal_key in ("projeto_id_vale", "logicalis_id", "ntt_id"):
                val = row.get(deal_key, "").strip()
                if val:
                    found = data_store.find_quote_by_deal_field(deal_key, val)
                    if found:
                        matched_id = found
                        break

        if matched_id:
            data_store.update_extraction_fields(matched_id, extract_fields, subject, user)
            data_store.save_annotation(matched_id, subject, ann_data, user)
            if deal_clean:
                data_store.save_deal(matched_id, subject, deal_clean, user)
            updated += 1
        else:
            new_id = _uuid.uuid4().hex[:12]
            quote  = {
                "id":             new_id,
                "is_manual":      True,
                "is_bulk_import": True,
                "date":           now_str,
                "from":           user,
                "products":       [],
                **extract_fields,
            }
            data_store.append_extraction_entry(quote, user)
            data_store.save_annotation(new_id, subject, ann_data, user)
            if deal_clean:
                data_store.save_deal(new_id, subject, deal_clean, user)
            created += 1

    session["import_result"] = {"created": created, "updated": updated, "errors": errors}
    return redirect(url_for("admin_import"))


# ── Bot API ──────────────────────────────────────────────────────────────────

def _api_key_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        configured = config.PORTAL_API_KEY
        if not configured:
            return jsonify({"error": "PORTAL_API_KEY not configured on server"}), 503
        if request.headers.get("X-API-Key", "") != configured:
            return jsonify({"error": "unauthorized"}), 401
        return f(*args, **kwargs)
    return decorated


@app.route("/api/v1/orders")
@_api_key_required
def api_orders():
    """Returns quotes that have an order_id set. Used by bot_to_ccw to know what to query."""
    quotes = data_store.load_extractions()
    deals  = data_store.load_deals()
    result = []
    for q in quotes:
        deal     = deals.get(q["id"], {})
        order_id = (deal.get("order_id") or "").strip()
        if order_id:
            result.append({
                "quote_id":  q["id"],
                "order_id":  order_id,
                "subject":   q.get("subject", ""),
                "last_sync": deal.get("last_ccw_sync", ""),
            })
    return jsonify(result)


@app.route("/api/v1/leadtime", methods=["POST"])
@_api_key_required
def api_leadtime():
    """Receives CCW XLS data from bot, detects scenario, computes lead_time, updates portal."""
    payload = request.get_json(silent=True)
    if not payload:
        return jsonify({"error": "invalid JSON"}), 400

    quote_id = payload.get("quote_id", "")
    order_id = payload.get("order_id", "")
    lines    = payload.get("lines", [])

    quote = next((q for q in data_store.load_extractions() if q["id"] == quote_id), None)
    if not quote:
        return jsonify({"error": "quote not found"}), 404

    result = data_store.process_ccw_sync(
        quote_id, quote.get("subject", ""), order_id, lines, "bot_ccw"
    )
    return jsonify({"ok": True, **result})


# ── Observabilidade ───────────────────────────────────────────────────────────

def _service_check(url: str, timeout: int = 4) -> dict:
    import urllib.request as _req
    import urllib.error   as _err2
    import time as _t
    t0 = _t.time()
    try:
        with _req.urlopen(url, timeout=timeout) as r:
            return {"status": "up", "latency_ms": round((_t.time() - t0) * 1000), "code": r.status}
    except _err2.HTTPError as e:
        lat = round((_t.time() - t0) * 1000)
        if e.code in (401, 403, 404):
            return {"status": "up", "latency_ms": lat, "code": e.code}
        return {"status": "degraded", "latency_ms": lat, "code": e.code}
    except Exception as ex:
        return {"status": "down", "latency_ms": None, "message": str(ex)[:80]}


@app.route("/admin/observability")
@admin_required
def admin_observability():
    return render_template("admin_observability.html")


@app.route("/api/v1/bot-status", methods=["POST"])
def api_bot_status_push():
    auth = request.headers.get("Authorization", "")
    if not config.PORTAL_API_KEY or auth != f"Bearer {config.PORTAL_API_KEY}":
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    data_store.save_bot_status(
        payload.get("runs", []),
        payload.get("order_errors", {}),
        payload.get("token_info", {}),
    )
    return jsonify({"ok": True})


@app.route("/api/v1/status")
@admin_required
def api_status():
    import time as _t
    import urllib.request as _req
    import urllib.error   as _err2

    services = {}

    # Portal — sempre up (estamos respondendo)
    services["portal"] = {"status": "up", "latency_ms": 0, "label": "Portal"}

    # Webhook server
    wh = _service_check(config.WEBHOOK_HEALTH_URL)
    wh["label"] = "Webhook Server"
    services["webhook"] = wh

    # Webex API
    webex_token = config.WEBEX_USER_TOKEN
    t0 = _t.time()
    try:
        r = _req.Request(
            "https://webexapis.com/v1/people/me",
            headers={"Authorization": f"Bearer {webex_token}" if webex_token else "Bearer probe"},
        )
        try:
            with _req.urlopen(r, timeout=5) as resp:
                lat = round((_t.time() - t0) * 1000)
                import json as _json2
                data = _json2.loads(resp.read())
                services["webex_api"] = {"status": "up", "latency_ms": lat, "token_valid": True, "user": data.get("displayName", "")}
        except _err2.HTTPError as e:
            lat = round((_t.time() - t0) * 1000)
            if e.code == 401:
                services["webex_api"] = {"status": "up", "latency_ms": lat, "token_valid": False, "token_expired": True}
            elif e.code == 403:
                services["webex_api"] = {"status": "up", "latency_ms": lat, "token_valid": True}
            else:
                services["webex_api"] = {"status": "degraded", "latency_ms": lat, "code": e.code}
    except Exception as ex:
        services["webex_api"] = {"status": "down", "latency_ms": None, "message": str(ex)[:60]}
    services["webex_api"]["label"] = "Webex API"

    # Bot CCW + token_info + order_errors — lidos do banco (push model)
    bs         = data_store.load_bot_status()
    bot_runs   = bs.get("runs") or []
    last_run   = bot_runs[-1] if bot_runs else None
    bot_st     = "nunca_executado"
    if last_run:
        bot_st = "ok" if last_run.get("failures", 0) == 0 else "falha_parcial"
        if last_run.get("success", 0) == 0:
            bot_st = "falha_total"
    services["bot_ccw"] = {
        "status":    bot_st,
        "label":     "Bot CCW (run_daily)",
        "last_run":  last_run.get("started_at") if last_run else None,
        "pushed_at": bs.get("pushed_at"),
    }

    # Contagem de emails recebidos (email_log table)
    email_count = 0
    try:
        with db.get_cursor() as cur:
            cur.execute("SELECT COUNT(*) AS n FROM email_log")
            email_count = (cur.fetchone() or {}).get("n", 0)
    except Exception:
        pass

    return jsonify({
        "services":     services,
        "token_info":   bs.get("token_info", {}),
        "email_count":  email_count,
        "bot_metrics":  {"runs": bot_runs},
        "order_errors": bs.get("order_errors", {}),
        "timestamp":    datetime.now().isoformat(),
    })


# ── Logs ──────────────────────────────────────────────────────────────────────

@app.route("/logs")
@login_required
def logs():
    all_logs    = data_store.load_audit_log()
    user_filter = request.args.get("user", "")
    all_users   = sorted({lg.get("user", "") for lg in all_logs})

    filtered = [lg for lg in all_logs if not user_filter or lg.get("user") == user_filter]
    return render_template("logs.html", logs=filtered, user_filter=user_filter, all_users=all_users)


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    data_store.ensure_default_user()
    print(f"Portal rodando em http://127.0.0.1:{config.PORTAL_PORT}")
    app.run(host="127.0.0.1", port=config.PORTAL_PORT, debug=False)
