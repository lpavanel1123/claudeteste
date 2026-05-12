import re
from pathlib import Path

_EMPTY = {
    "products": [],
    "requester_name_xls": "NA",
    "department_xls": "NA",
    "cnpj_xls": "NA",
    "project_ref": "NA",
}


def read_xls(filepath: str) -> dict:
    try:
        suffix = Path(filepath).suffix.lower()
        if suffix == ".xlsx":
            return _read_xlsx(filepath)
        return _read_xls_legacy(filepath)
    except Exception as exc:
        print(f"  [xls_reader] erro ao ler {filepath}: {exc}")
        return _EMPTY.copy()


# ── readers ──────────────────────────────────────────────────────────────────

def _read_xls_legacy(filepath: str) -> dict:
    import xlrd
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    rows = [[_cell_val(ws.cell(r, c)) for c in range(ws.ncols)] for r in range(ws.nrows)]
    return _parse_rows(rows)


def _read_xlsx(filepath: str) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = [[_clean(cell.value) for cell in row] for row in ws.iter_rows()]
    return _parse_rows(rows)


# ── parser ────────────────────────────────────────────────────────────────────

def _parse_rows(rows: list) -> dict:
    result = _EMPTY.copy()
    result["products"] = _extract_products(rows)
    result.update(_extract_header_fields(rows))
    return result


def _extract_products(rows: list) -> list:
    """Find product table by locating a header row with Part Number / Qtd."""
    header_idx = None
    qty_col = part_col = desc_col = None

    for i, row in enumerate(rows):
        cells = [str(c).lower() for c in row]
        if any("part" in c for c in cells) and any(re.search(r"qtd|qty|quant", c) for c in cells):
            header_idx = i
            for j, c in enumerate(cells):
                if re.search(r"qtd|qty|quant", c):
                    qty_col = j
                elif "part" in c:
                    part_col = j
                elif re.search(r"descri|desc", c):
                    desc_col = j
            break

    if header_idx is None:
        return []

    products = []
    for row in rows[header_idx + 1:]:
        qty = row[qty_col] if qty_col is not None else ""
        pn = row[part_col] if part_col is not None else ""
        desc = row[desc_col] if desc_col is not None else ""
        if not qty and not pn:
            continue
        if str(qty).strip() or str(pn).strip():
            products.append({
                "qty": _qty(qty),
                "part_number": str(pn).strip(),
                "description": str(desc).strip() if desc else "NA",
            })
    return products


def _extract_header_fields(rows: list) -> dict:
    """Extract requester, department, CNPJ and project ref from header area."""
    data = {
        "requester_name_xls": "NA",
        "department_xls": "NA",
        "cnpj_xls": "NA",
        "project_ref": "NA",
    }
    name_labels = re.compile(r"solicitante|requisitante|nome|requester", re.IGNORECASE)
    dept_labels = re.compile(r"departamento|depto|department|setor", re.IGNORECASE)
    cnpj_labels = re.compile(r"cnpj", re.IGNORECASE)
    ref_labels  = re.compile(r"ref\.?\s*projeto|proj.*ref|project ref", re.IGNORECASE)

    for row in rows:
        flat = " ".join(str(c) for c in row)
        for j, cell in enumerate(row):
            cell_str = str(cell).strip()
            # label in one cell, value in next
            next_val = str(row[j + 1]).strip() if j + 1 < len(row) else ""

            if name_labels.search(cell_str) and next_val and next_val != "NA":
                data["requester_name_xls"] = next_val
            elif dept_labels.search(cell_str) and next_val and next_val != "NA":
                data["department_xls"] = next_val
            elif cnpj_labels.search(cell_str) and next_val and next_val != "NA":
                data["cnpj_xls"] = next_val
            elif ref_labels.search(cell_str) and next_val and next_val != "NA":
                data["project_ref"] = next_val

        # Also check if the entire row flat contains a CNPJ pattern
        if data["cnpj_xls"] == "NA":
            m = re.search(r"\d{2}[\.\s]?\d{3}[\.\s]?\d{3}[\s/]?\d{4}[-\s]?\d{2}", flat)
            if m:
                data["cnpj_xls"] = m.group(0).strip()

    return data


# ── utils ─────────────────────────────────────────────────────────────────────

def _cell_val(cell) -> str:
    import xlrd
    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return ""
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        v = cell.value
        return str(int(v)) if v == int(v) else str(v)
    return str(cell.value).strip()


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _qty(value) -> str:
    try:
        f = float(str(value).replace(",", "."))
        return str(int(f)) if f == int(f) else str(f)
    except (ValueError, TypeError):
        return str(value).strip()
